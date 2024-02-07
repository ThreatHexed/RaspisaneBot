from openpyxl import load_workbook
import os
from database import DataBase

def __init__(self) -> None:
        self.worksheet = load_workbook(os.getcwd() + r'/documents/raspisanie.xlsx')['Лист1']

def update_shelude(self):
    self.worksheet = load_workbook(os.getcwd() + r'/documents/raspisanie.xlsx')['Лист1']

def get_timetable(self, weekday, user_id):
        group = DataBase().get_group(user_id)

        timetable = ''
        
        for column in range(3, 23):
            if str(self.worksheet.cell(10,column).value) in group:
                user_column = column

        weekdays = ['понедельник', 'вторник', 'среда', "четверг", "пятница", "суббота", "воскресенье"]

        for i in range(14, 79, 11):
            
            weekday_timetable = weekdays.index((self.worksheet.cell(i,1).value.strip(' ')).lower()) + 1
            if weekday == weekday_timetable:
                timetable = f'Расписание на {(self.worksheet.cell(i,1).value).lower()} для группы {group}'
                for id, x in enumerate(range(i, i + 10)):
                    if (id+1) % 2 == 1:
                        if id == 6:
                            # timetable += f'\n   {worksheet.cell(x,user_column).value}'
                            pass
                        else:
                            if id == 8:
                                timetable += f'\n4. {self.worksheet.cell(x,user_column).value}'
                            else: timetable += f'\n{round((id)/2)+1}. {self.worksheet.cell(x,user_column).value}'

                    else:
                        if id == 7:
                            # timetable += f'\n   {worksheet.cell(x,user_column).value}'
                            pass
                        else: timetable += f'\n   {self.worksheet.cell(x,user_column).value}'
        timetable = timetable.replace("None", '---------------')
        return timetable

