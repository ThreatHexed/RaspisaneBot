import pandas as pd
import json
from datetime import date
from datetime import datetime
from openpyxl import load_workbook
import os
from database import DataBase

def para_prepod(prepod):

    try:

        weekday = ['0', "Понедельник", "Вторник", "Среда", "Четверг",
                    "Пятница", "Суббота", "Понедельник", "Понедельник"]

        if datetime.now().hour > 8:
            userdate = str(date.today().isoweekday()+1)

        else:
            userdate = str(date.today().isoweekday())

        with open("2.json", "r", encoding='utf-8') as f:
            data = json.load(f)

        day = data[userdate]

        excel = pd.read_excel(io='test1.xlsx',
            engine='openpyxl',
            usecols='A:S',
            header=8,
            )

        para = f'{weekday[int(userdate)]}\n\n'

        for id, day_number in enumerate(day, 2):
            hui = False
            for group in excel:
                
                if str(excel[group].loc[day[day_number]]).find(str(prepod)) > -1 and hui == False:
                    hui = True
                    para += (f"____________\n{excel[group].loc[0]} \n {excel[group].loc[day[day_number]-1]} \n {excel[group].loc[day[day_number]]}\n")

                elif str(excel[group].loc[day[day_number]]).find(str(prepod)) > -1 and hui == True:
                    para += (f"{excel[group].loc[0]} \n {excel[group].loc[day[day_number]-1]} \n {excel[group].loc[day[day_number]]}\n")

        if prepod == 'Колесников':
            para += "\nТакого преподавателя я в хуй не дул, и вообще, дядя, кто ты нахуй такой"
        print(para)
        return para
    except Exception as e:
        return f"Ошибки \n {e} \n А теперь пиздуй чинить"
    
def get_timetable(weekday, user_id):
    group = DataBase.get_group(user_id)

    timetable = ''
    workbook = load_workbook(os.getcwd() + '/raspisanie.xlsx')
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    
    for column in range(3, 23):
        if str(worksheet.cell(10,column).value) in group:
            user_column = column

    weekdays = ['понедельник', 'вторник', 'среда', "четверг", "пятница", "суббота", "воскресенье"]

    for i in range(14, 66, 9):
        
        weekday_timetable = weekdays.index((worksheet.cell(i,1).value.strip(' ')).lower()) + 1

        if weekday == weekday_timetable:
            timetable = f'Расписание на {(worksheet.cell(i,1).value).lower()} для группы {group}'
            for id, x in enumerate(range(i, i + 8)):
                if (id+1) % 2 == 1:
                    timetable += f'\n{round((id)/2)+1}. {worksheet.cell(x,user_column).value}'
                else:
                    timetable += f'\n   {worksheet.cell(x,user_column).value}'
    timetable = timetable.replace("None", '---------------')
    print(type(timetable))
    return timetable
